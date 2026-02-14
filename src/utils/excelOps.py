import os
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

DEBUGMODE = False
STANDARD_COL_LIST = [""]
COLOR_RED = "FF0000"
COLOR_YELLOW = "FFFF00"
FONT_FORMAT_HEADER_BOLD = Font(
    name="Calibri",
    bold=True,
    italic=False,
    size=14,
)
 
CELL_STYLE_YELLOW = PatternFill(
    patternType="solid",
    fgColor=COLOR_YELLOW,
)
FONT_FORMAT_BOLD_RED = Font(
    underline=None,
    color=COLOR_RED,
    bold=True,
    italic=False,
)


FONT_FORMAT_HEADER_BOLD_RED = Font(
    name="Calibri",
    bold=True,
    italic=False,
    size=14,
    color="FF0000",
)

FONT_FORMAT_HEADER_BOLD_WHITE = Font(
    name="Calibri",
    bold=True,
    italic=False,
    size=12,
    color="FFFFFF",
)

CELL_STYLE_HEADER_FILL = PatternFill(
    fill_type="solid",
    start_color="00338D",
    end_color="00338D",
)


class ExcelOps:
    """ "
    Excel class for simpler excel manipulation
    """

    def __init__(
        self,
        filepath: str,
        create_new: bool = False,
        sheets: list = None,
    ) -> None:
        """Inititalise an excel object

        Args:
            filepath: filepath to excel sheet
            create_new: if create_new is selected then it will create directory and create a new file
            sheets: create a series of excel sheets
        
        Returns:
            None
        """

        if not isinstance(sheets, list):
            return "Sheets should be a list of string"

        if not create_new and os.path.isfile(filepath):
            raise Exception(f"File {filepath} does not exists.")

        if create_new:
            if os.path.isfile(filepath):
                os.remove(filepath)
            if not os.path.isdir(os.path.dirname(filepath)):
                os.mkdir(os.path.dirname(filepath))
            openpyxl.Workbook().save(filepath)

        self.wb = load_workbook(filepath)
        self.filepath = filepath

        if sheets is not None:
            self.wb.worksheets[0].title = "Sheet_delete"
            for sheet in sheets:
                self.wb.create_sheet(sheet)
            self.wb.remove(self.wb["Sheet_delete"])

    def __del__(self):
        """Destructor function
            Close excel file
        Args: None
        
        Returns: None
        """
        if hasattr(self, "wb") and self.wb is not None:
            try:
                self.wb.close()
            except Exception as e:
                print(f"Error closing workbook: {e}")

    def __str__(self) -> str:
        print(f"filepath = {self.filepath}")


    def print_header(
        self,
        headers=None,
        sheetname=None,
        sheetindex=None,
        insertrows=True,
    ) -> None:
        """Print Header
        Args:
            headers     : List of headers of a workpaper
            sheetname   : Select sheet by name
            sheetindex  : Select sheet by index
            insertrows : Insert row if neeeded

        Return: 
            None
        """

        if headers is None or not isinstance(headers, list) or not headers:
            print("No Headers passed in are written")
            return None

        try:
            if sheetname is None and sheetindex is None:
                curr_sheet = self.wb.worksheets[0]
            elif sheetname is not None:
                curr_sheet = self.wb[sheetname]
            elif sheetindex is not None:
                curr_sheet = self.wb.worksheets[0]
        except Exception as e:
            print(f"Error: Sheet not found - {e}")

        if insertrows:
            curr_sheet.insert_rows(1, len(headers) + 2)

        for header in range(len(headers)):
            curr_sheet[f"A{str(header+1)}"].value = headers[header]
            curr_sheet[f"A{str(header+1)}"].font = FONT_FORMAT_HEADER_BOLD

        self.wb.save(self.filepath)

    def print_from_df(
        self,
        df,
        sheetname=None,
        sheetindex=None,
    ) -> None:
        """Print from dataframe
        Args:
            df          : Dataframe with the content
            sheetname   : Select sheet by name
            sheetindex  : Select sheet by index

        Return:
            None
        """
        if not isinstance(df, pd.DataFrame):
            raise ("Error - Dataframe is required")

        try:
            if sheetname is None and sheetindex is None:
                curr_sheet = self.wb.worksheets[0]
            elif sheetname is not None:
                curr_sheet = self.wb[sheetname]
            elif sheetindex is not None:
                curr_sheet = self.wb.worksheets[0]
        except Exception as e:
            print(f"Error: Sheet not found - {e}")

        for r in dataframe_to_rows(df, index=False, header=True):
            curr_sheet.append(r)


        for cell in curr_sheet[1]:
            cell.fill = CELL_STYLE_HEADER_FILL
            cell.font = FONT_FORMAT_HEADER_BOLD_WHITE

        for row in curr_sheet:
            for cell in row:
                if isinstance(cell.value, str):
                    val = cell.value.strip().lower()
                    if val.lower() == 'true':
                        cell.value = True
                    elif val.lower() == 'false':
                        cell.value = False

        self._auto_adjust_column_width(curr_sheet)
        self.wb.save(self.filepath)
        return


    def rename_sheet(
        self,
        sheetname_old:str,
        sheetname_new:str,
    ) -> None:
        """Rename Sheet
        Args:
            sheetname_old: old sheetname
            sheetname_new: new sheetname

        Return:
            None
        """
        try:
            self.wb[sheetname_old].title = sheetname_new
            self.wb.save(self.filepath)
        except:
            print(f"Error Renaming Sheet {sheetname_old}")
        return

    def add_sheet(
        self,
        sheetname:str,
    ) -> None:
        """Add Sheet
        Args:
            sheetname: new sheet

        Return:
            None
        """
        
        try:
            self.wb.create_sheet(sheetname)
            self.wb.save(self.filepath)
        except:
            print(f"Error adding Sheet {sheetname}")
        return


    def del_sheet(
        self,
        sheetname:str,
    ) -> None:
        """Delete Sheet
        Args:
            sheetname: delete sheet

        Return:
            None
        """
        try:
            self.wb.remove(self.wb[sheetname])
            self.wb.save(self.filepath)
        except:
            print(f"Error adding Sheet {sheetname}")
        return

    def highlight_cell(
        self,
        search_term: str,
        sheetname=None,
        sheetindex=None,
    ) -> bool:
        """Highlight cell that contains exact match of the search term
        Parameters:
            filepath: of file
            search_term:  term to find in Excel

        Description:
            Simple Function to find exact match of cells
            that contains the search terms
            Possible Development:
                SUBSTRING matching
                REGEX matching
        """
        try:
            if sheetname is None and sheetindex is None:
                curr_sheet = self.wb[0]
            elif sheetname is not None:
                curr_sheet = self.wb[sheetname]
            elif sheetindex is not None:
                curr_sheet = self.wb[sheetindex]
        except Exception as e:
            raise (f"Error: Sheet not found - {e}")

        for row in range(1, curr_sheet.max_row + 1):
            for column in range(1, curr_sheet.max_column + 1):
                cell = curr_sheet.cell(row, column)
                if  search_term.lower() == str(cell.value).lower():
                    cell.fill = CELL_STYLE_YELLOW
                    cell.font = FONT_FORMAT_BOLD_RED

        self.wb.save(self.filepath)

    def add_border(self, cell, position, borderType):
        return

    def _auto_adjust_column_width(
        self,
        sheet:str,
    ) -> None:
        """Adjust column width
        Parameters:
            sheet: Sheet object that needs to be columns adjusted

        Description:
            Adjust column width to prevent cut-off
        """
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    cell_value = str(cell.value)
                    max_length = max(max_length, len(cell_value))

                except Exception:
                    pass

            adjusted_width = max_length + 5
            sheet.column_dimensions[column_letter].width = adjusted_width
